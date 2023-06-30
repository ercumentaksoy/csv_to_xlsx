# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import pandas as pd
import numpy as np
import openpyxl


data = pd.read_csv('D:/myphd/groundstation/2020/Ankara__Ankara_Bahçelievler__2020_01_01_to_2021_01_01.csv')

wb = openpyxl.Workbook()
sayfa = wb.active

a2 = len(data)          ### toplam satır sayısı
a3 = len(data.columns)  ### toplam sütun sayısı
print('satır uzunluğu:', a2)
print('sütun sayısı:', a3)

for x in range(a3):  ### sütun başlıklarını yazdırma dongüsü
    c = x + 1
    sayfa.cell(row = 1, column = c).value = data.columns[x]
    
for x in range(a2): ### tüm satırlardaki verileri excele yazdırma döngüsü
    for y in range(a3):
        r = x + 2
        c = y + 1
        sayfa.cell(row = r, column = c).value = data.iat[x,y]
        
wb.save("D:/myphd/groundstation/2020/excel/Ankara__Ankara_Bahçelievler__2020_01_01_to_2021_01_01.xlsx")
print('işlem başarıyla tamamlandı. Excel dosyanız oluşturuldu')
